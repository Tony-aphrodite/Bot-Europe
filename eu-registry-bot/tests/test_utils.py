"""
Tests for utility modules
"""

import os
import time
import pytest
import tempfile
from unittest.mock import MagicMock, patch

from src.utils.retry import (
    retry,
    RetryExecutor,
    RetryStrategy,
    CircuitBreaker,
    calculate_delay,
)
from src.utils.state import (
    StateManager,
    SubmissionState,
    SubmissionStep,
)
from src.utils.captcha import (
    CaptchaDetector,
    CaptchaHandler,
    CaptchaType,
    CaptchaDetectionResult,
)


class TestRetryDecorator:
    """Tests for retry decorator."""

    def test_retry_success_first_attempt(self):
        """Test function succeeds on first attempt."""
        call_count = 0

        @retry(max_attempts=3, delay=0.1)
        def successful_func():
            nonlocal call_count
            call_count += 1
            return "success"

        result = successful_func()
        assert result == "success"
        assert call_count == 1

    def test_retry_success_after_failures(self):
        """Test function succeeds after some failures."""
        call_count = 0

        @retry(max_attempts=3, delay=0.1, exceptions=(ValueError,))
        def flaky_func():
            nonlocal call_count
            call_count += 1
            if call_count < 3:
                raise ValueError("Temporary error")
            return "success"

        result = flaky_func()
        assert result == "success"
        assert call_count == 3

    def test_retry_all_attempts_fail(self):
        """Test all attempts fail."""
        call_count = 0

        @retry(max_attempts=3, delay=0.1, exceptions=(ValueError,))
        def always_fails():
            nonlocal call_count
            call_count += 1
            raise ValueError("Always fails")

        with pytest.raises(ValueError):
            always_fails()

        assert call_count == 3

    def test_retry_non_retryable_exception(self):
        """Test non-retryable exception is raised immediately."""
        call_count = 0

        @retry(max_attempts=3, delay=0.1, exceptions=(ValueError,))
        def raises_other():
            nonlocal call_count
            call_count += 1
            raise TypeError("Not retryable")

        with pytest.raises(TypeError):
            raises_other()

        assert call_count == 1


class TestCalculateDelay:
    """Tests for delay calculation."""

    def test_fixed_delay(self):
        """Test fixed delay strategy."""
        assert calculate_delay(1, 5.0, RetryStrategy.FIXED) == 5.0
        assert calculate_delay(2, 5.0, RetryStrategy.FIXED) == 5.0
        assert calculate_delay(3, 5.0, RetryStrategy.FIXED) == 5.0

    def test_linear_delay(self):
        """Test linear delay strategy."""
        assert calculate_delay(1, 2.0, RetryStrategy.LINEAR) == 2.0
        assert calculate_delay(2, 2.0, RetryStrategy.LINEAR) == 4.0
        assert calculate_delay(3, 2.0, RetryStrategy.LINEAR) == 6.0

    def test_exponential_delay(self):
        """Test exponential delay strategy."""
        assert calculate_delay(1, 1.0, RetryStrategy.EXPONENTIAL) == 1.0
        assert calculate_delay(2, 1.0, RetryStrategy.EXPONENTIAL) == 2.0
        assert calculate_delay(3, 1.0, RetryStrategy.EXPONENTIAL) == 4.0

    def test_max_delay_cap(self):
        """Test maximum delay cap."""
        assert calculate_delay(10, 10.0, RetryStrategy.EXPONENTIAL, max_delay=30.0) == 30.0


class TestRetryExecutor:
    """Tests for RetryExecutor class."""

    def test_execute_success(self):
        """Test successful execution."""
        executor = RetryExecutor(max_attempts=3, base_delay=0.1)

        def success_func():
            return "result"

        result = executor.execute(success_func)
        assert result == "result"

    def test_execute_with_callbacks(self):
        """Test execution with callbacks."""
        executor = RetryExecutor(max_attempts=3, base_delay=0.1)

        success_called = False
        retry_count = 0

        def on_success(result, attempt):
            nonlocal success_called
            success_called = True

        def on_retry(attempt, error):
            nonlocal retry_count
            retry_count += 1

        executor.on_success(on_success)
        executor.on_retry(on_retry)

        call_count = 0

        def flaky_func():
            nonlocal call_count
            call_count += 1
            if call_count < 2:
                raise ConnectionError("Temporary")
            return "ok"

        result = executor.execute(flaky_func)
        assert result == "ok"
        assert success_called
        assert retry_count == 1


class TestCircuitBreaker:
    """Tests for CircuitBreaker."""

    def test_circuit_closed_initially(self):
        """Test circuit is closed initially."""
        cb = CircuitBreaker(failure_threshold=3)
        assert not cb.is_open

    def test_circuit_opens_after_threshold(self):
        """Test circuit opens after failure threshold."""
        cb = CircuitBreaker(failure_threshold=3, recovery_timeout=60.0)

        for _ in range(3):
            cb.record_failure()

        assert cb.is_open

    def test_circuit_closes_after_success(self):
        """Test circuit closes after success."""
        cb = CircuitBreaker(failure_threshold=2)
        cb.record_failure()
        cb.record_failure()
        assert cb.is_open

        # Mock time to simulate recovery timeout
        cb._last_failure_time = time.time() - 61
        assert not cb.is_open  # Now half-open

        cb.record_success()
        assert not cb.is_open


class TestStateManager:
    """Tests for StateManager."""

    def test_create_and_load_state(self):
        """Test creating and loading state."""
        with tempfile.TemporaryDirectory() as tmpdir:
            manager = StateManager(tmpdir)

            state = manager.create_state(
                application_id="app123",
                country="portugal",
                portal="gov.pt",
            )

            assert state.application_id == "app123"
            assert state.country == "portugal"
            assert state.current_step == SubmissionStep.INITIALIZED

            # Load and verify
            loaded = manager.load_state(state.submission_id)
            assert loaded is not None
            assert loaded.application_id == state.application_id

    def test_update_step(self):
        """Test updating submission step."""
        with tempfile.TemporaryDirectory() as tmpdir:
            manager = StateManager(tmpdir)

            state = manager.create_state(
                application_id="app456",
                country="france",
                portal="service-public.fr",
            )

            manager.update_step(state, SubmissionStep.AUTHENTICATED)
            assert state.current_step == SubmissionStep.AUTHENTICATED

            # Reload and verify
            loaded = manager.load_state(state.submission_id)
            assert loaded.current_step == SubmissionStep.AUTHENTICATED

    def test_mark_failed(self):
        """Test marking submission as failed."""
        with tempfile.TemporaryDirectory() as tmpdir:
            manager = StateManager(tmpdir)

            state = manager.create_state(
                application_id="app789",
                country="portugal",
                portal="gov.pt",
            )

            manager.mark_failed(state, "Connection timeout", "Server not responding")

            assert state.current_step == SubmissionStep.FAILED
            assert state.error_message == "Connection timeout"

    def test_get_incomplete_submissions(self):
        """Test getting incomplete submissions."""
        with tempfile.TemporaryDirectory() as tmpdir:
            manager = StateManager(tmpdir)

            # Create some states
            state1 = manager.create_state("app1", "portugal", "gov.pt")
            state2 = manager.create_state("app2", "france", "service-public.fr")
            state3 = manager.create_state("app3", "portugal", "gov.pt")

            # Complete one, fail another
            manager.update_step(state1, SubmissionStep.COMPLETED)
            manager.mark_failed(state2, "Error")

            incomplete = manager.get_incomplete_submissions()

            # Only state3 should be incomplete
            assert len(incomplete) == 1
            assert incomplete[0].application_id == "app3"

    def test_recoverable_step(self):
        """Test determining recoverable step."""
        with tempfile.TemporaryDirectory() as tmpdir:
            manager = StateManager(tmpdir)

            state = manager.create_state("app", "portugal", "gov.pt")

            # Test different steps
            state.current_step = SubmissionStep.AUTHENTICATED
            assert manager.get_recoverable_step(state) == SubmissionStep.AUTHENTICATED

            state.current_step = SubmissionStep.SUBMITTED
            assert manager.get_recoverable_step(state) is None  # Can't recover

            state.current_step = SubmissionStep.FAILED
            assert manager.get_recoverable_step(state) == SubmissionStep.INITIALIZED


class TestCaptchaDetector:
    """Tests for CaptchaDetector."""

    def test_no_captcha_detected(self):
        """Test when no CAPTCHA is present."""
        mock_driver = MagicMock()
        mock_driver.find_element.side_effect = Exception("Not found")

        detector = CaptchaDetector(mock_driver)
        result = detector.detect()

        assert not result.detected
        assert result.captcha_type == CaptchaType.UNKNOWN

    def test_recaptcha_detected(self):
        """Test reCAPTCHA detection."""
        mock_driver = MagicMock()
        mock_element = MagicMock()
        mock_element.is_displayed.return_value = True
        mock_element.get_attribute.return_value = "https://recaptcha.net/..."

        def find_element_side_effect(by, selector):
            if "recaptcha" in selector:
                return mock_element
            raise Exception("Not found")

        mock_driver.find_element.side_effect = find_element_side_effect

        detector = CaptchaDetector(mock_driver)
        result = detector.detect()

        assert result.detected
        assert result.captcha_type == CaptchaType.RECAPTCHA_V2


class TestCaptchaHandler:
    """Tests for CaptchaHandler."""

    def test_no_captcha_returns_true(self):
        """Test returns True when no CAPTCHA."""
        mock_driver = MagicMock()
        mock_driver.find_element.side_effect = Exception("Not found")

        handler = CaptchaHandler(mock_driver, auto_wait=False)
        result = handler.check_and_handle()

        assert result

    def test_notification_callback_called(self):
        """Test notification callback is called on CAPTCHA detection."""
        mock_driver = MagicMock()
        mock_element = MagicMock()
        mock_element.is_displayed.return_value = True

        def find_element_side_effect(by, selector):
            if "recaptcha" in selector:
                return mock_element
            raise Exception("Not found")

        mock_driver.find_element.side_effect = find_element_side_effect

        handler = CaptchaHandler(mock_driver, auto_wait=False)

        notification_received = []

        def callback(msg):
            notification_received.append(msg)

        handler.set_notification_callback(callback)
        handler.check_and_handle()

        assert len(notification_received) == 1
        assert "CAPTCHA detected" in notification_received[0]


class TestExcelReader:
    """Tests for ExcelReader (requires openpyxl)."""

    @pytest.fixture
    def sample_excel(self, tmp_path):
        """Create a sample Excel file for testing."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "Municipalities"

        # Headers
        ws['A1'] = 'Code'
        ws['B1'] = 'Municipality'
        ws['C1'] = 'Province'
        ws['D1'] = 'Population'
        ws['E1'] = 'Status'

        # Data rows
        data = [
            ('L01001', 'Test City 1', 'Province A', 50000, 'pending'),
            ('L01002', 'Test City 2', 'Province A', 30000, 'pending'),
            ('L01003', 'Test City 3', 'Province B', 75000, 'completed'),
            ('L01004', 'Test City 4', 'Province B', 20000, 'pending'),
        ]

        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        file_path = tmp_path / "test_municipalities.xlsx"
        wb.save(file_path)
        wb.close()

        return str(file_path)

    def test_excel_reader_basic(self, sample_excel):
        """Test basic Excel reading functionality."""
        try:
            from src.utils.excel_reader import ExcelReader, MunicipalityRecord
        except ImportError:
            pytest.skip("openpyxl not installed")

        reader = ExcelReader(
            sample_excel,
            custom_mappings={
                'code': 'code',
                'municipality': 'name',
                'province': 'province',
                'population': 'population_total',
                'status': 'status',
            }
        )

        records = reader.read_all()
        reader.close()

        assert len(records) == 4
        assert records[0].code == 'L01001'
        assert records[0].name == 'Test City 1'
        assert records[0].province == 'Province A'
        assert records[0].population_total == 50000

    def test_excel_reader_generator(self, sample_excel):
        """Test generator-based reading for memory efficiency."""
        try:
            from src.utils.excel_reader import ExcelReader
        except ImportError:
            pytest.skip("openpyxl not installed")

        with ExcelReader(
            sample_excel,
            custom_mappings={
                'code': 'code',
                'municipality': 'name',
            }
        ) as reader:
            count = 0
            for record in reader.read_generator():
                count += 1
                assert record.name is not None

            assert count == 4

    def test_excel_reader_summary(self, sample_excel):
        """Test getting file summary."""
        try:
            from src.utils.excel_reader import ExcelReader
        except ImportError:
            pytest.skip("openpyxl not installed")

        reader = ExcelReader(
            sample_excel,
            custom_mappings={
                'code': 'code',
                'municipality': 'name',
            }
        )

        summary = reader.get_summary()
        reader.close()

        assert summary['data_rows'] == 4
        assert 'code' in summary['mapped_columns']
        assert 'name' in summary['mapped_columns']

    def test_excel_reader_file_not_found(self):
        """Test handling of missing file."""
        try:
            from src.utils.excel_reader import ExcelReader
        except ImportError:
            pytest.skip("openpyxl not installed")

        reader = ExcelReader("/nonexistent/file.xlsx")

        with pytest.raises(FileNotFoundError):
            reader.read_all()


class TestBatchProcessor:
    """Tests for BatchProcessor."""

    @pytest.fixture
    def sample_records(self):
        """Create sample MunicipalityRecord objects."""
        try:
            from src.utils.excel_reader import MunicipalityRecord
        except ImportError:
            pytest.skip("openpyxl not installed")

        return [
            MunicipalityRecord(code='L01', name='City 1', status='pending', row_number=2),
            MunicipalityRecord(code='L02', name='City 2', status='pending', row_number=3),
            MunicipalityRecord(code='L03', name='City 3', status='completed', row_number=4),
        ]

    def test_batch_processor_success(self, sample_records):
        """Test successful batch processing."""
        try:
            from src.utils.excel_reader import BatchProcessor, ExcelBatchResult
        except ImportError:
            pytest.skip("openpyxl not installed")

        # Mock ExcelReader
        mock_reader = MagicMock()
        mock_reader.read_all.return_value = sample_records

        processed = []

        def process_func(record):
            processed.append(record.name)

        processor = BatchProcessor(
            mock_reader,
            process_func,
            skip_statuses=['completed'],
        )

        result = processor.run()

        assert result.total_records == 3
        assert result.successful == 2
        assert result.skipped == 1
        assert len(processed) == 2
        assert 'City 1' in processed
        assert 'City 2' in processed

    def test_batch_processor_with_errors(self, sample_records):
        """Test batch processing with failures."""
        try:
            from src.utils.excel_reader import BatchProcessor
        except ImportError:
            pytest.skip("openpyxl not installed")

        mock_reader = MagicMock()
        mock_reader.read_all.return_value = sample_records[:2]  # Only pending ones

        def process_func(record):
            if record.code == 'L02':
                raise RuntimeError("Test error")

        errors_received = []

        def on_error(record, error):
            errors_received.append((record.code, str(error)))

        processor = BatchProcessor(
            mock_reader,
            process_func,
            on_error=on_error,
        )

        result = processor.run()

        assert result.successful == 1
        assert result.failed == 1
        assert len(errors_received) == 1
        assert errors_received[0][0] == 'L02'


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
