"""
Data file batch processing module for municipality lists.
Supports: .xlsx, .xls, .csv, .docx
"""

import os
import csv
import zipfile
import xml.etree.ElementTree as ET
from typing import List, Dict, Any, Optional, Generator
from dataclasses import dataclass, field
from enum import Enum

from ..core.logger import setup_logger

logger = setup_logger(__name__)

# Try to import openpyxl, fall back gracefully if not available
try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel (.xlsx) support disabled.")


class ExcelColumnMapping(Enum):
    """Standard column mappings for municipality Excel files."""
    MUNICIPALITY_CODE = "code"
    MUNICIPALITY_NAME = "name"
    PROVINCE = "province"
    REGION = "region"
    POPULATION_TOTAL = "population_total"
    POPULATION_URBAN = "population_urban"
    POPULATION_RURAL = "population_rural"
    STATUS = "status"
    NOTES = "notes"


@dataclass
class MunicipalityRecord:
    """Single municipality record from Excel."""
    code: str
    name: str
    province: Optional[str] = None
    region: Optional[str] = None
    population_total: Optional[int] = None
    population_urban: Optional[int] = None
    population_rural: Optional[int] = None
    status: str = "pending"
    notes: Optional[str] = None
    row_number: int = 0
    extra_data: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "code": self.code,
            "name": self.name,
            "province": self.province,
            "region": self.region,
            "population_total": self.population_total,
            "population_urban": self.population_urban,
            "population_rural": self.population_rural,
            "status": self.status,
            "notes": self.notes,
            "row_number": self.row_number,
            "extra_data": self.extra_data,
        }


@dataclass
class ExcelBatchResult:
    """Result of batch processing."""
    total_records: int
    successful: int
    failed: int
    skipped: int
    records: List[MunicipalityRecord]
    errors: List[Dict[str, Any]] = field(default_factory=list)

    @property
    def success_rate(self) -> float:
        """Calculate success rate as percentage."""
        if self.total_records == 0:
            return 0.0
        return (self.successful / self.total_records) * 100


class ExcelReader:
    """
    Reads municipality lists from Excel files for batch processing.
    """

    # Default column name mappings (case-insensitive)
    DEFAULT_COLUMN_MAPPINGS = {
        # Spanish
        "concejo": ExcelColumnMapping.MUNICIPALITY_NAME,
        "codigo": ExcelColumnMapping.MUNICIPALITY_CODE,
        "código": ExcelColumnMapping.MUNICIPALITY_CODE,
        "provincia": ExcelColumnMapping.PROVINCE,
        "region": ExcelColumnMapping.REGION,
        "región": ExcelColumnMapping.REGION,
        "habitantes": ExcelColumnMapping.POPULATION_TOTAL,
        "hab_total": ExcelColumnMapping.POPULATION_TOTAL,
        "nº total hab": ExcelColumnMapping.POPULATION_TOTAL,
        "hab_urbano": ExcelColumnMapping.POPULATION_URBAN,
        "nº hab urbano": ExcelColumnMapping.POPULATION_URBAN,
        "hab_rural": ExcelColumnMapping.POPULATION_RURAL,
        "nº hab rural": ExcelColumnMapping.POPULATION_RURAL,
        "estado": ExcelColumnMapping.STATUS,
        "notas": ExcelColumnMapping.NOTES,
        # Portuguese
        "concelho": ExcelColumnMapping.MUNICIPALITY_NAME,
        "município": ExcelColumnMapping.MUNICIPALITY_NAME,
        "municipio": ExcelColumnMapping.MUNICIPALITY_NAME,
        "distrito": ExcelColumnMapping.PROVINCE,
        "população": ExcelColumnMapping.POPULATION_TOTAL,
        "populacao": ExcelColumnMapping.POPULATION_TOTAL,
        # French
        "commune": ExcelColumnMapping.MUNICIPALITY_NAME,
        "département": ExcelColumnMapping.PROVINCE,
        "departement": ExcelColumnMapping.PROVINCE,
        "population": ExcelColumnMapping.POPULATION_TOTAL,
        # English
        "municipality": ExcelColumnMapping.MUNICIPALITY_NAME,
        "name": ExcelColumnMapping.MUNICIPALITY_NAME,
        "code": ExcelColumnMapping.MUNICIPALITY_CODE,
        "province": ExcelColumnMapping.PROVINCE,
        "status": ExcelColumnMapping.STATUS,
        "notes": ExcelColumnMapping.NOTES,
    }

    def __init__(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        header_row: int = 1,
        custom_mappings: Optional[Dict[str, str]] = None,
    ):
        """
        Initialize Excel reader.

        Args:
            file_path: Path to Excel file (.xlsx, .xls)
            sheet_name: Specific sheet to read (default: first sheet)
            header_row: Row number containing headers (1-indexed)
            custom_mappings: Custom column name to field mappings
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel support. "
                "Install with: pip install openpyxl"
            )

        self.file_path = file_path
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.custom_mappings = custom_mappings or {}
        self._workbook = None
        self._sheet = None
        self._column_map: Dict[int, ExcelColumnMapping] = {}

    def _validate_file(self) -> None:
        """Validate the Excel file exists and is readable."""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        ext = os.path.splitext(self.file_path)[1].lower()
        if ext not in ('.xlsx', '.xls', '.xlsm'):
            raise ValueError(f"Unsupported file format: {ext}. Use .xlsx, .xls, .xlsm, or use DataReader for other formats")

    def _load_workbook(self) -> None:
        """Load the Excel workbook."""
        self._validate_file()

        logger.info(f"Loading Excel file: {self.file_path}")
        self._workbook = load_workbook(self.file_path, read_only=True, data_only=True)

        if self.sheet_name:
            if self.sheet_name not in self._workbook.sheetnames:
                raise ValueError(
                    f"Sheet '{self.sheet_name}' not found. "
                    f"Available sheets: {self._workbook.sheetnames}"
                )
            self._sheet = self._workbook[self.sheet_name]
        else:
            self._sheet = self._workbook.active

        logger.info(f"Loaded sheet: {self._sheet.title}")

    def _detect_columns(self) -> None:
        """Detect and map columns based on header row."""
        if not self._sheet:
            raise RuntimeError("Workbook not loaded")

        self._column_map = {}

        for col_idx, cell in enumerate(
            self._sheet.iter_rows(min_row=self.header_row, max_row=self.header_row).__next__(),
            start=1
        ):
            if cell.value is None:
                continue

            header = str(cell.value).strip().lower()

            # Check custom mappings first
            if header in self.custom_mappings:
                field_name = self.custom_mappings[header]
                try:
                    mapping = ExcelColumnMapping(field_name)
                    self._column_map[col_idx] = mapping
                    logger.debug(f"Column {col_idx} '{header}' -> {mapping.value} (custom)")
                    continue
                except ValueError:
                    pass

            # Check default mappings
            if header in self.DEFAULT_COLUMN_MAPPINGS:
                mapping = self.DEFAULT_COLUMN_MAPPINGS[header]
                self._column_map[col_idx] = mapping
                logger.debug(f"Column {col_idx} '{header}' -> {mapping.value}")

        # Validate required columns
        required = {ExcelColumnMapping.MUNICIPALITY_NAME}
        found = set(self._column_map.values())

        if not required.issubset(found):
            missing = required - found
            raise ValueError(
                f"Required columns not found: {[m.value for m in missing]}. "
                f"Found columns: {[m.value for m in found]}"
            )

        logger.info(f"Detected {len(self._column_map)} mapped columns")

    def _parse_row(self, row, row_number: int) -> Optional[MunicipalityRecord]:
        """Parse a single row into a MunicipalityRecord."""
        data = {}
        extra = {}

        for col_idx, cell in enumerate(row, start=1):
            value = cell.value

            if col_idx in self._column_map:
                mapping = self._column_map[col_idx]
                if mapping == ExcelColumnMapping.MUNICIPALITY_CODE:
                    data["code"] = str(value).strip() if value else ""
                elif mapping == ExcelColumnMapping.MUNICIPALITY_NAME:
                    data["name"] = str(value).strip() if value else ""
                elif mapping == ExcelColumnMapping.PROVINCE:
                    data["province"] = str(value).strip() if value else None
                elif mapping == ExcelColumnMapping.REGION:
                    data["region"] = str(value).strip() if value else None
                elif mapping == ExcelColumnMapping.POPULATION_TOTAL:
                    data["population_total"] = int(value) if value else None
                elif mapping == ExcelColumnMapping.POPULATION_URBAN:
                    data["population_urban"] = int(value) if value else None
                elif mapping == ExcelColumnMapping.POPULATION_RURAL:
                    data["population_rural"] = int(value) if value else None
                elif mapping == ExcelColumnMapping.STATUS:
                    data["status"] = str(value).strip().lower() if value else "pending"
                elif mapping == ExcelColumnMapping.NOTES:
                    data["notes"] = str(value).strip() if value else None
            elif value is not None:
                # Store unmapped columns in extra_data
                extra[f"col_{col_idx}"] = value

        # Skip empty rows
        if not data.get("name"):
            return None

        # Generate code if not provided
        if not data.get("code"):
            data["code"] = f"AUTO_{row_number}"

        return MunicipalityRecord(
            row_number=row_number,
            extra_data=extra,
            **data
        )

    def read_all(self) -> List[MunicipalityRecord]:
        """
        Read all municipality records from the Excel file.

        Returns:
            List of MunicipalityRecord objects
        """
        self._load_workbook()
        self._detect_columns()

        records = []
        start_row = self.header_row + 1

        for row_idx, row in enumerate(
            self._sheet.iter_rows(min_row=start_row),
            start=start_row
        ):
            try:
                record = self._parse_row(row, row_idx)
                if record:
                    records.append(record)
            except Exception as e:
                logger.warning(f"Error parsing row {row_idx}: {e}")

        logger.info(f"Read {len(records)} municipality records")
        return records

    def read_generator(self) -> Generator[MunicipalityRecord, None, None]:
        """
        Generator that yields municipality records one at a time.
        Memory-efficient for large files.

        Yields:
            MunicipalityRecord objects
        """
        self._load_workbook()
        self._detect_columns()

        start_row = self.header_row + 1

        for row_idx, row in enumerate(
            self._sheet.iter_rows(min_row=start_row),
            start=start_row
        ):
            try:
                record = self._parse_row(row, row_idx)
                if record:
                    yield record
            except Exception as e:
                logger.warning(f"Error parsing row {row_idx}: {e}")

    def get_summary(self) -> Dict[str, Any]:
        """
        Get summary information about the Excel file.

        Returns:
            Dictionary with file summary
        """
        self._load_workbook()
        self._detect_columns()

        total_rows = self._sheet.max_row - self.header_row
        mapped_columns = [m.value for m in self._column_map.values()]

        return {
            "file_path": self.file_path,
            "sheet_name": self._sheet.title,
            "total_sheets": len(self._workbook.sheetnames),
            "available_sheets": self._workbook.sheetnames,
            "header_row": self.header_row,
            "data_rows": total_rows,
            "mapped_columns": mapped_columns,
            "total_columns": self._sheet.max_column,
        }

    def close(self) -> None:
        """Close the workbook and release resources."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None
            self._sheet = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


class BatchProcessor:
    """
    Processes multiple municipalities from Excel in batch.
    """

    def __init__(
        self,
        excel_reader: ExcelReader,
        process_func,
        on_progress=None,
        on_error=None,
        skip_statuses: Optional[List[str]] = None,
    ):
        """
        Initialize batch processor.

        Args:
            excel_reader: ExcelReader instance
            process_func: Function to process each record (receives MunicipalityRecord)
            on_progress: Callback for progress updates (current, total, record)
            on_error: Callback for errors (record, error)
            skip_statuses: List of statuses to skip (e.g., ["completed", "skipped"])
        """
        self.reader = excel_reader
        self.process_func = process_func
        self.on_progress = on_progress
        self.on_error = on_error
        self.skip_statuses = skip_statuses or ["completed", "success"]

    def run(self) -> ExcelBatchResult:
        """
        Run batch processing on all records.

        Returns:
            ExcelBatchResult with processing statistics
        """
        records = self.reader.read_all()
        total = len(records)

        result = ExcelBatchResult(
            total_records=total,
            successful=0,
            failed=0,
            skipped=0,
            records=records,
        )

        for idx, record in enumerate(records, start=1):
            # Check if should skip
            if record.status.lower() in self.skip_statuses:
                result.skipped += 1
                record.status = "skipped"
                logger.info(f"Skipping {record.name} (status: {record.status})")
                continue

            # Progress callback
            if self.on_progress:
                self.on_progress(idx, total, record)

            try:
                # Process the record
                self.process_func(record)
                record.status = "completed"
                result.successful += 1
                logger.info(f"Processed {record.name} successfully")

            except Exception as e:
                record.status = "failed"
                record.notes = str(e)
                result.failed += 1
                result.errors.append({
                    "record": record.to_dict(),
                    "error": str(e),
                })

                logger.error(f"Failed to process {record.name}: {e}")

                if self.on_error:
                    self.on_error(record, e)

        logger.info(
            f"Batch complete: {result.successful} success, "
            f"{result.failed} failed, {result.skipped} skipped"
        )

        return result


class CSVReader:
    """
    Reads municipality lists from CSV files.
    """

    def __init__(
        self,
        file_path: str,
        delimiter: str = ',',
        encoding: str = 'utf-8',
        custom_mappings: Optional[Dict[str, str]] = None,
    ):
        self.file_path = file_path
        self.delimiter = delimiter
        self.encoding = encoding
        self.custom_mappings = custom_mappings or {}
        self._column_map: Dict[int, ExcelColumnMapping] = {}

    def _detect_delimiter(self) -> str:
        """Auto-detect CSV delimiter."""
        with open(self.file_path, 'r', encoding=self.encoding) as f:
            first_line = f.readline()
            if ';' in first_line and ',' not in first_line:
                return ';'
            elif '\t' in first_line:
                return '\t'
            return ','

    def read_all(self) -> List[MunicipalityRecord]:
        """Read all records from CSV."""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"CSV file not found: {self.file_path}")

        logger.info(f"Loading CSV file: {self.file_path}")

        # Auto-detect delimiter
        delimiter = self._detect_delimiter()

        records = []
        with open(self.file_path, 'r', encoding=self.encoding, errors='replace') as f:
            reader = csv.DictReader(f, delimiter=delimiter)

            for row_idx, row in enumerate(reader, start=2):
                try:
                    # Map columns
                    data = {}
                    for key, value in row.items():
                        if key is None:
                            continue
                        key_lower = key.strip().lower()

                        # Check custom mappings
                        if key_lower in self.custom_mappings:
                            field_name = self.custom_mappings[key_lower]
                            data[field_name] = value.strip() if value else ""
                        # Check default mappings
                        elif key_lower in ExcelReader.DEFAULT_COLUMN_MAPPINGS:
                            mapping = ExcelReader.DEFAULT_COLUMN_MAPPINGS[key_lower]
                            data[mapping.value] = value.strip() if value else ""

                    # Create record if we have a name
                    name = data.get('name', '')
                    if name:
                        record = MunicipalityRecord(
                            code=data.get('code', f'CSV_{row_idx}'),
                            name=name,
                            province=data.get('province'),
                            status=data.get('status', 'pending'),
                            row_number=row_idx,
                        )
                        # Parse population if present
                        if 'population_total' in data and data['population_total']:
                            try:
                                record.population_total = int(data['population_total'])
                            except ValueError:
                                pass
                        records.append(record)

                except Exception as e:
                    logger.warning(f"Error parsing CSV row {row_idx}: {e}")

        logger.info(f"Read {len(records)} records from CSV")
        return records

    def get_summary(self) -> Dict[str, Any]:
        """Get CSV file summary."""
        records = self.read_all()
        return {
            "file_path": self.file_path,
            "data_rows": len(records),
            "format": "CSV",
        }

    def close(self):
        pass

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass


class DocxReader:
    """
    Reads municipality lists from Word documents (.docx).
    Parses tables or structured text with municipality data.
    """

    def __init__(
        self,
        file_path: str,
        custom_mappings: Optional[Dict[str, str]] = None,
    ):
        self.file_path = file_path
        self.custom_mappings = custom_mappings or {}

    def _extract_text(self) -> List[str]:
        """Extract text paragraphs from docx."""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"DOCX file not found: {self.file_path}")

        with zipfile.ZipFile(self.file_path, 'r') as zip_ref:
            xml_content = zip_ref.read('word/document.xml')

        root = ET.fromstring(xml_content)
        ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}

        paragraphs = []
        for p in root.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p'):
            text_parts = []
            for t in p.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t'):
                if t.text:
                    text_parts.append(t.text)
            if text_parts:
                text = ''.join(text_parts).strip()
                if text:
                    paragraphs.append(text)

        return paragraphs

    def read_all(self) -> List[MunicipalityRecord]:
        """
        Read municipality data from docx.
        Expects format: Municipality name followed by a number (count/population).
        """
        logger.info(f"Loading DOCX file: {self.file_path}")
        paragraphs = self._extract_text()

        records = []
        current_province = None
        i = 0

        while i < len(paragraphs):
            text = paragraphs[i]

            # Skip header lines
            if any(skip in text.lower() for skip in [
                'equipos de', 'total de', 'municipio', 'total equipos',
                'header', 'título', 'title'
            ]):
                i += 1
                continue

            # Check for province/region header (e.g., "Anexo:Municipios de Álava")
            if text.startswith("Anexo:"):
                current_province = text.replace("Anexo:Municipios de ", "").replace("Anexo: Municipios de ", "").strip()
                i += 1
                continue

            # Try to parse as municipality (name followed by number)
            if i + 1 < len(paragraphs):
                name = text
                try:
                    count = int(paragraphs[i + 1])
                    record = MunicipalityRecord(
                        code=f"DOCX_{len(records) + 1:05d}",
                        name=name,
                        province=current_province,
                        population_total=count,
                        status="pending",
                        row_number=len(records) + 1,
                    )
                    records.append(record)
                    i += 2
                    continue
                except ValueError:
                    pass

            i += 1

        logger.info(f"Read {len(records)} records from DOCX")
        return records

    def get_summary(self) -> Dict[str, Any]:
        """Get DOCX file summary."""
        records = self.read_all()
        return {
            "file_path": self.file_path,
            "data_rows": len(records),
            "format": "DOCX",
        }

    def close(self):
        pass

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass


class DataReader:
    """
    Unified reader that automatically detects file format.
    Supports: .xlsx, .xls, .csv, .docx
    """

    SUPPORTED_FORMATS = {
        '.xlsx': 'excel',
        '.xls': 'excel',
        '.xlsm': 'excel',
        '.csv': 'csv',
        '.docx': 'docx',
    }

    def __init__(
        self,
        file_path: str,
        custom_mappings: Optional[Dict[str, str]] = None,
        **kwargs
    ):
        """
        Initialize unified data reader.

        Args:
            file_path: Path to data file
            custom_mappings: Column name to field mappings
            **kwargs: Additional arguments for specific readers
        """
        self.file_path = file_path
        self.custom_mappings = custom_mappings or {}
        self.kwargs = kwargs
        self._reader = None
        self._format = None

        self._detect_format()
        self._create_reader()

    def _detect_format(self):
        """Detect file format from extension."""
        ext = os.path.splitext(self.file_path)[1].lower()

        if ext not in self.SUPPORTED_FORMATS:
            raise ValueError(
                f"Unsupported file format: {ext}. "
                f"Supported formats: {list(self.SUPPORTED_FORMATS.keys())}"
            )

        self._format = self.SUPPORTED_FORMATS[ext]
        logger.info(f"Detected format: {self._format} for {self.file_path}")

    def _create_reader(self):
        """Create appropriate reader based on format."""
        if self._format == 'excel':
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl required for Excel files. Install: pip install openpyxl")
            self._reader = ExcelReader(
                self.file_path,
                custom_mappings=self.custom_mappings,
                **{k: v for k, v in self.kwargs.items() if k in ['sheet_name', 'header_row']}
            )
        elif self._format == 'csv':
            self._reader = CSVReader(
                self.file_path,
                custom_mappings=self.custom_mappings,
                **{k: v for k, v in self.kwargs.items() if k in ['delimiter', 'encoding']}
            )
        elif self._format == 'docx':
            self._reader = DocxReader(
                self.file_path,
                custom_mappings=self.custom_mappings,
            )

    def read_all(self) -> List[MunicipalityRecord]:
        """Read all records from the file."""
        return self._reader.read_all()

    def get_summary(self) -> Dict[str, Any]:
        """Get file summary."""
        summary = self._reader.get_summary()
        summary['detected_format'] = self._format
        return summary

    def close(self):
        """Close the reader."""
        if self._reader:
            self._reader.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    @staticmethod
    def get_supported_formats() -> List[str]:
        """Get list of supported file extensions."""
        return list(DataReader.SUPPORTED_FORMATS.keys())
